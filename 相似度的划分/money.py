from openpyxl import load_workbook
import re

# 万元 美元 元 千元 万余元 余万元 多万元

# 汇率转换
万 = 10000
千 = 1000
美元 = 6.8
英镑 = 8.6
日元 = 0.06
港元 = 0.88
韩元 = 0.006

database = "E:\data_analysis\判决书分词\贪污受贿罪_组合.xlsx"

def money_sim(compare_num):
    wb = load_workbook(database)
    sheet = wb.worksheets[1]
    length = len(sheet["B"])
    list_value = []
    for i in range(length-1):
        value_of_block = str(sheet.cell(row=i + 2, column=5).value)
        # 提取带小数点的数字
        value = float(re.findall(r"\d+\.?\d*", value_of_block)[0])
        judge1 = re.search("万", value_of_block)
        judge2 = re.search("千", value_of_block)
        judge3 = re.search("美", value_of_block)
        judge4 = re.search("英镑", value_of_block)
        judge5 = re.search("日元", value_of_block)
        judge6 = re.search("港元", value_of_block)
        judge7 = re.search("韩元", value_of_block)
        if judge1 is not None:
            value *= 万
        elif judge2 is not None:
            value *= 千
        if judge3 is not None:
            value *= 美元
        elif judge4 is not None:
            value *= 英镑
        elif judge5 is not None:
            value *= 日元
        elif judge6 is not None:
            value *= 港元
        elif judge7 is not None:
            value *= 韩元
        list_value.append(value)
    compare_value = list_value[int(compare_num)-1]
    list_value_abs = [abs(value-compare_value) for value in list_value]
    serial_num = list(range(len(sheet["B"])-1))
    serial_num = [value + 2 for value in serial_num]
    dictionary = list(zip(serial_num, list_value_abs, list_value))
    dictionary = sorted(dictionary, key=lambda x:x[1])
    print("The row number of the text equals", compare_num+1,", the compare value euqals", compare_value)
    print("The top 10 simularity as follow:")
    print("row num, difference, value")
    for i in range(10):
        print(dictionary[i+1])


if __name__ == '__main__':
    num = input("Please input the row number of text you want to compare.")
    money_sim(int(num)-1)

