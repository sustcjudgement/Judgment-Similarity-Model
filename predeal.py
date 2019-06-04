# -*- coding:utf-8 -*-
import simhash
import openpyxl
import threading
import simple_tfidf
import re
import synonyms

content_list = []
for i in range(1, 150):
    simple_tfidf.convertencoding('lib\\text{}.txt'.format(i))
    file = open('lib\\text{}.txt'.format(i), 'rb')
    content = file.read().decode()
    content_list.append(content)
    file.close()

# 将案件文本全部提取出来,并做处理
filename = r'C:\Users\kingway\Desktop\判决书分词\贪污受贿罪_组合.xlsx'
case_base = openpyxl.load_workbook(filename)
ws = case_base.worksheets[1]
row = ws.max_row  # 行数
column = ws.max_column  # 列数

file2 = open('预处理_考虑词性2.txt', 'w')


def delete(string):
    '''

    :param string: 要处理的字符串
    :return:
    '''
    new_string = ''
    for i in range(len(string)):
        if u'\u9fa5' >= string[i] >= u'\u4e00':  # 判断是否是中文
            continue
        new_string += string[i]
    return new_string


def predeal(pos, next):
    for i in range(pos, next):
        print(i)
        # mon=0
        no = ws.cell(row=i, column=1).value
        cont = ws.cell(row=i, column=2).value
        money = ws.cell(row=i, column=5).value
        type1 = ws.cell(row=i, column=3).value
        type2 = ws.cell(row=i, column=4).value
        # print(money)
        if re.search('万', money):
            mon = delete(money)
            mon = float(mon) * 10000
        elif re.search('千', money):
            mon = delete(money)
            mon = float(mon) * 1000
        else:
            mon = delete(money)
            mon = float(mon)

            # try:
            #     mon = float(mon)
            # except ValueError:
            #     print(mon)
        money = mon
        hashcode = simhash.simhash(cont, content_list, type=True)
        if hashcode.__str__() == '00':
            continue
        keywords = ''
        for tu in hashcode.tf_idf[:21]:
            keywords += str(tu[0]).replace('\'', '') + ' '
        ws.cell(row=i, column=7, value=keywords)
        print(keywords)
        file2.write(no + ' ' + type1 + ' ' + type2 + ' ' + str(money) + ' ' + keywords + '\n')


pos = 2
size = 2500
threads = []

for i in range(5):
    nex = size + pos if size + pos <= row else row
    t = threading.Thread(target=predeal, args=(pos, nex + 1))
    threads.append(t)
    t.start()
    pos = nex + 1

for i in threads:
    i.join()

predeal(2, row)
case_base.save(filename)
case_base.close()
file2.close()
