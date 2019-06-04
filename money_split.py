import re
import openpyxl
import threading

filename = r'C:\Users\kingway\Desktop\判决书分词\贪污受贿罪_组合_end.xlsx'
case_base = openpyxl.load_workbook(filename)
ws = case_base.worksheets[1]
row = ws.max_row
column = ws.max_column

file2 = open('money_result.txt', 'w')


def delete(string):
    new_string = ''
    for i in range(len(string)):
        if u'\u9fa5' >= string[i] >= u'\u4e00':  # 判断是否是中文
            continue
        new_string += string[i]
    return new_string


def predeal(pos, next):
    for i in range(pos, next):
        no = ws.cell(row=i, column=1).value
        content = ws.cell(row=i, column=3).value
        if re.search('万', content):
            money = delete(content)
            money = float(money) * 10000
        elif re.search('千', content):
            money = delete(content)
            money = float(money) * 1000
        else:
            money = delete(content)
            money = float(money)
        content = money
        file2.write(no + ' ' + str(content) + '\n')


pos = 30
size = 2100
threads = []

for i in range(4):
    nex = size + pos if size + pos <= row else row
    t = threading.Thread(target=predeal, args=(pos, nex + 1))
    threads.append(t)
    pos = nex + 1

for i in threads:
    i.start()
