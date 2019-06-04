from sklearn.cluster import KMeans
import numpy as np
import openpyxl

file = '贪污受贿罪_一审_提取.xlsx'

if __name__=="__main__":
    wb=openpyxl.load_workbook(file)
    ws=wb.worksheets[0]
    col=ws.max_column
    row = ws.max_row
    data=[]
    data_index=[]
    for i in range(1,row+1):
        data_index.append(int(ws.cell(row=i,column=1).value))
        data.append(list(str(ws.cell(row=i,column=3).value)))
    X=np.array(data)
    kmeans = KMeans().fit(X)
    label = list(kmeans.labels_)
    num=np.zeros(8,dtype=int)
    for i in label:
        num[i]+=1
    print(num)
    # for i in range(1,row+1):
    #     ws.cell(row=i,column=4).value=label[i-1]
    # wb.save(filename='贪污受贿罪_一审_提取.xlsx')

