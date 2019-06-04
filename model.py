import simhash
import openpyxl
import simple_tfidf
import math
import project1
import re
import split
import time
import gensim

filename = r'C:\Users\kingway\Desktop\判决书分词\判决书数据库\贪污受贿罪_一审.xlsx'
case_lis = r'C:\Users\kingway\Desktop\判决书分词\预处理_考虑词性.txt'


# 获取计算idf值时的文档
def get_idf_content(iter_range=150):
    content_list = []
    for i in range(1, iter_range + 1):
        simple_tfidf.convertencoding('lib\\text{}.txt'.format(i))
        file = open('lib\\text{}.txt'.format(i), 'rb')
        content = file.read().decode()
        content_list.append(content)
        file.close()
    return content_list


# 打开文件并返回,用于测试
def source_file():
    cases = openpyxl.load_workbook(filename)
    predeal = open(case_lis, 'r')
    return cases, predeal


# 打开预处理后的结果储存文件
def source_file_stan():
    predeal = open(case_lis, 'r')
    return predeal


# 关闭打开的文件
def source_close(filelist):
    for i in filelist:
        i.close()


# 返回元组的第三个元素
def key(it):
    return it[1]


# 金额比较
def compare_money(money, data, number):
    sim = []
    for m in data:
        has = m.replace('\n', '')
        da = has.split(' ')
        da[-1] = da[-1].replace('\n', '')
        tupl = (da[0], math.fabs(money - float(da[3])), da[4:len(da) - 1])
        sim.append(tupl)
    sim.sort(key=key)
    num = 1
    # print('与本案涉案金额相似的案件有：')
    result = []
    for tu in sim:
        if num > number:
            break
        # print(tu[0], tu[1])
        result.append(tu)
        num += 1
    return result


# 去除字符串中的汉字
def delete(string):
    new_string = ''
    for i in range(len(string)):
        if u'\u9fa5' >= string[i] >= u'\u4e00':  # 判断是否是中文
            continue
        new_string += string[i]
    return new_string


# 将金额文本转化为数字
def money_convert(money):
    if re.search('万', money):
        mon = delete(money)
        mon = float(mon) * 10000
    elif re.search('千', money):
        mon = delete(money)
        mon = float(mon) * 1000
    else:
        mon = delete(money)
        mon = float(mon)
    return mon


# 利用已有判决书进行计算
def precess(money_dis=False, number=500):
    file = []
    cases, predeal = source_file()
    file.append(cases)
    file.append(predeal)
    ws = cases.worksheets[1]
    data = predeal.readlines()
    count = 1
    content_list = get_idf_content()
    for j in range(40, 100):
        if count > 4:
            break
        name = ws.cell(row=j, column=10).value  # 获取被告人姓名
        value = ws.cell(row=j, column=22).value  # 获取庭审过程
        if len(name.split('、')) > 1:  # 处理只有一名被告人的文本
            continue
        result = split.split_text(name, value)
        ca = simhash.simhash(result, content_list, type=True)
        if len(ca.__str__()) < 64:
            continue
        simi = {}
        if money_dis:
            money = project1.distinguish(value)
            money = money_convert(money)
            sim_case = compare_money(float(money), data, number)
            for has in sim_case:
                # start = time.time()
                simi[has[0]] = ca.matrix_dis(has[2])
                # print("total compare time ",time.time()-start)
        else:
            for has in data:
                has = has.strip('\n')
                da = has.split(' ')
                if len(da[-1])<64:
                    continue
                # if len(da[4:]) < 10:
                #     continue
                # start = time.time()
                # simila = ca.matrix_dis(da[4:len(da) - 1])
                # print(time.time()-start)
                simi[da[0]] = ca.count_cos(da[-1])
        simi = sorted(simi.items(), key=lambda x: x[1], reverse=True)
        num = 1
        print('与第%d号判决书相似的案件有：' % j)
        for tu in simi:
            if num > 10:
                break
            if int(tu[0]) == j:
                continue
            print(tu)
            num += 1
        count += 1
    source_close(file)


# 实际使用时调用的函数
def similarity(name, content):
    '''

    :param name: 被告人姓名
    :param content:庭审过程
    :return:
    '''
    file = []
    predeal = source_file_stan()
    file.append(predeal)
    data = predeal.readlines()
    content_list = get_idf_content()
    result, money = split.split_text(name, content)
    ca = simhash.simhash(result, content_list)
    print(ca.tf_idf)
    simi = {}
    for has in data:
        has = has.strip('\n')
        da = has.split(' ')
        if len(da[4:]) < 10:
            continue
        simila = ca.matrix_dis(da[4:len(da) - 1])
        simi[da[0]] = simila
    simi = sorted(simi.items(), key=lambda x: x[1], reverse=True)
    num = 1
    print('与该案件相似判决书有：')
    for tu in simi:
        if num > 10:
            break
        print(tu)
        num += 1
    source_close(file)


# 测试函数
def test():
    file = []
    cases, predeal = source_file()
    file.append(cases)
    file.append(predeal)
    ws = cases.worksheets[1]
    data = predeal.readlines()
    content_list = get_idf_content()
    candidate = [22, 23, 30, 36, 37, 44, 45, 46, 48, 714, 2, 185, 315, 71, 80]
    name = ws.cell(row=15, column=10).value  # 获取被告人姓名
    value = ws.cell(row=15, column=22).value  # 获取庭审过程
    result = split.split_text(name, value)
    ca = simhash.simhash(result, content_list, type=True, )
    sim = {}
    for has in data:
        has = has.strip('\n')
        da = has.split(' ')
        if int(da[0]) not in candidate:
            continue
        simila = ca.matrix_dis(da[4:len(da) - 1])
        sim[da[0]] = simila
    sim = sorted(sim.items(), key=lambda x: x[1], reverse=True)
    print('与第%d号判决书相似的案件有：' % 15)
    for tu in sim:
        print(tu)
    source_close(file)


# def test2():
    # import gensim
#     file = []
#     cases, predeal = source_file()
#     file.append(cases)
#     file.append(predeal)
#     ws = cases.worksheets[1]
#     data = predeal.readlines()
#     content_list = get_idf_content()
#     W2V = gensim.models.Word2Vec.load('word2vec')
#     candidate = [22, 23, 30, 36, 37, 44, 45, 46, 48, 714, 2, 185, 315, 71, 80]
#     name = ws.cell(row=15, column=10).value  # 获取被告人姓名
#     value = ws.cell(row=15, column=22).value  # 获取庭审过程
#     result = split.split_text(name, value)
#     ca = simhash.simhash(result, content_list, type=True, )
#     sim = {}
#     for has in data:
#         has = has.strip('\n')
#         da = has.split(' ')
#         if int(da[0]) not in candidate:
#             continue
#         print(int(da[0]))
#         simila = ca.matrix_dis_self(da[4:len(da) - 1], W2V)
#         sim[da[0]] = simila
#     sim = sorted(sim.items(), key=lambda x: x[1], reverse=True)
#     print('与第%d号判决书相似的案件有：' % 15)
#     for tu in sim:
#         print(tu)
#     source_close(file)


if __name__ == '__main__':
    start = time.time()
    precess(money_dis=False)
    # test2()
    print(time.time() - start)
