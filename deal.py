import openpyxl
import re

filename = r'C:\Users\kingway\Desktop\判决书分词\贪污受贿罪.xlsx'

file = openpyxl.load_workbook(filename)

ws = file.worksheets[1]
print("open load_file")
file_write = openpyxl.Workbook()
write = file_write.create_sheet()
print("open write_file")
row = ws.max_row
column = ws.max_column
print(row, column)
o = 1
for i in range(1, column + 1):
    cell_value = ws.cell(row=o, column=i).value
    write.cell(row=o, column=i).value = cell_value
    print(ws.cell(row=o, column=i).value)
o += 1

for i in range(2, row + 1):
    value = ws.cell(row=i, column=4).value
    judge = False
    if re.match(r'二审', str(value)):
        judge = True
    # if re.match(r"受贿罪", str(value)):
    #     judge = True
    if judge:
        print(value)
        for j in range(1, column + 1):
            cell_value = ws.cell(row=i, column=j).value
            write.cell(row=o, column=j).value = cell_value
        o += 1

file_write.save(filename="贪污受贿罪_二审.xlsx")
