from openpyxl import load_workbook
from openpyxl import Workbook
import re

wb = load_workbook("贪污受贿罪_组合.xlsx")
sheet = wb.worksheets[1]
length = len(sheet["B"])
wsht = Workbook()
wsheet = wsht.active
wsheet.title = "词跨度精简结果"
# get each line of xlsx
for i in range(length):
    # get the value of B0 to Bn
    value_of_block = str(sheet.cell(row=i + 1, column=2).value)
    list = ["人民检察院nt", "刑诉v", "起诉书n", "本院n", "提起公诉v", "依法n", "合议庭n", "开庭审理n", "检察员n", "出庭n",
            "诉讼vn", "审理vn", "指控vn", "辩护人n","组成v","适用v","立案n","分院n","指派n","受理v","开庭n","律师n",
            "事务所律师n","审查vn","上述事实n","进行v","检vn","刑n","生于v","决定v","参加v","利用n","担任v","亚力n","提交v"
            ,"现tg","羁押于n","看守所n","检诉v","讨论v","决定v","宝刑n","息刑n","初字n","简易程序n","担任v","申张v","统n"
            ,"出v","娥n","经n","乙n","拥军v","终结v","过程n"]
    l_of_list = len(list)
    # remove the useless word
    for j in range(l_of_list):
        judge = re.search(list[j], value_of_block)
        if judge is not None:
            # print(list[j])
            value_of_block = value_of_block.replace(list[j], "")
        penny = re.search("补助款v", value_of_block)
        if penny is not None:
            value_of_block = value_of_block.replace("补助款v", "补助款n")
        penny = re.search("补助费v", value_of_block)
        if penny is not None:
            value_of_block = value_of_block.replace("补助费v", "补助费n")
    value_of_block = value_of_block.strip()
    if i == 0:
        wsheet.cell(row=i + 1, column=1).value = "位置"
        wsheet.cell(row=i + 1, column=2).value = "精简结果"
        wsheet.cell(row=i + 1, column=3).value = "金额"
        wsheet.cell(row=i + 1, column=4).value = "贪污"
        wsheet.cell(row=i + 1, column=5).value = "受贿"
    else:
        wsheet.cell(row=i + 1, column=1).value = sheet.cell(row=i + 1, column=1).value
        # print(value_of_block)
        wsheet.cell(row=i + 1, column=2).value = value_of_block
        wsheet.cell(row=i + 1, column=3).value = sheet.cell(row=i + 1, column=3).value
        wsheet.cell(row=i + 1, column=4).value = sheet.cell(row=i + 1, column=4).value
        wsheet.cell(row=i + 1, column=5).value = sheet.cell(row=i + 1, column=5).value

wsht.save("词跨度精简结果.xlsx")
print("finish")