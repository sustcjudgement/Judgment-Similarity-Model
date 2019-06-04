# encoding=utf-8
import jieba
import jieba.posseg as psg
from openpyxl import load_workbook


# wb = load_workbook("贪污受贿罪.xlsx")
# sheet = wb.worksheets[1]
# value_of_block1 = str(sheet.cell(row=7526, column=22).value)
# value_of_block2 = str(sheet.cell(row=1505, column=22).value)
# value_of_block3 = str(sheet.cell(row=6527, column=22).value)
# value_of_block4 = str(sheet.cell(row=8002, column=22).value)
# value_of_block5 = str(sheet.cell(row=4500, column=22).value)



# 对新判决文本的处理
def split_text(text):
    jieba.load_userdict("family.txt")
    jieba.load_userdict("finance.txt")
    jieba.load_userdict("law.txt")
    jieba.load_userdict("place.txt")
    '''

    :param name: 被告名字
    :param text: 庭审过程
    :return: 返回分词结果
    '''
    global result  # 与被告有关
    global money_result  # 收受 受贿
    global sum_money_result  # 收受总数
    global get_money_result  # 骗取  贪污
    global get_sum_money_result  # 收受总数

    result = ""
    money_result = ""
    get_money_result = ""
    sum_money_result = ""
    get_sum_money_result = ""
    text_point = text
    text1 = jieba.cut(text)

    # text = str(",".join(text1)).replace("，、", "")
    # text = str(text).replace("，,", "")
    # text = str(text).replace("、,", "")
    # text = str(text).replace("。,", "")
    # text = str(text).replace("；,", "")
    # text = str(text).replace("：", "")
    # text = str(text).replace(".", "")
    #
    # text = jieba.cut(text)
    print("导入词库后的结果",", ".join(text1))

str = "李某某拟好补偿协议"
textorig = jieba.cut(str)
print("原始结果", ", ".join(textorig))
textmodi = split_text(str)
